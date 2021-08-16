import requests
import re
import sys
from openpyxl import Workbook, load_workbook
from lxml import etree
# 对表格的操作
wb = Workbook()
ws1 = wb.create_sheet("Mysheet")           #创建一个sheet
line = 2
ws1["A1"] = "domain"
ws1["B1"] = "title"
ws1["C1"] = "ip"
arg1 = int(sys.argv[1])
arg2 = int(sys.argv[2])

for ip1 in range(arg1, arg2):
    for ip2 in range(1,256):
        url = "http://gcsec.top:81/Less-1/?name=210.44.{}.{}".format(ip1,ip2)
        print(url)
        try:
            text = requests.get(url,timeout=3).content.decode()
        except:
            pass
        else:
            ret1 = re.findall('= (.+)\.',text)
            print(ret1)
            for i in ret1:
                try:
                    u = "http://"+i
                    r1 = requests.get(u, timeout=3)
                    html = r1.content.decode()
                    title = re.findall('<title>(.+)</title>',html)
                except:
                    ws1["A" + str(line)] = "http://"+i
                    ws1["B" + str(line)] = "no"
                    ws1["C" + str(line)] = "210.44.{}.{}".format(ip1,ip2)
                    line += 1
                else:
                    ws1["A" + str(line)] = "http://" + i
                    ws1["B" + str(line)] = str(title)
                    ws1["C" + str(line)] = "210.44.{}.{}".format(ip1,ip2)
                    line += 1

    wb.save("{}.xlsx".format(ip1))

